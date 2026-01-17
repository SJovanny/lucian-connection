import json
from openpyxl import load_workbook

# Configuration des fichiers
INPUT_FILE = 'export_produits_2026-01-16 10_46_46_696a0906ea8e6.xlsx'
OUTPUT_FILE = 'import_produits.sql'

# Mapping des noms de catégories du CSV vers les slugs de votre BDD
CATEGORY_MAPPING = {
    'ASSAISONNEMENT': 'assaisonnement',
    'ALIMENTAIRE': 'alimentaire',
    'DIVERS': 'divers',
    'BOISSONS SANS ALCOOL': 'boissons-sans-alcool',
    'BOISSONS ALCOOLISÉES': 'boissons-alcoolisees',
    'SEAMOSS': 'seamoss'
}

def clean_text(text):
    """Nettoie les problèmes d'encodage courants (ex: AFRICAN€™S -> AFRICAN'S)"""
    if not text: return ""
    return text.replace("€™", "'").replace("'", "''") # Échappement pour SQL

def create_slug(name, product_id):
    """Crée un slug URL-friendly"""
    safe_name = "".join([c if c.isalnum() else "-" for c in name.lower()])
    return f"{safe_name}-{product_id}"

sql_statements = []

# En-tête du fichier SQL
sql_statements.append("-- Importation massive des produits")
sql_statements.append("BEGIN;")

try:
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    # Sauter la ligne d'en-tête
    for row in rows[1:]:
        # Vérification basique pour éviter les lignes vides
        if len(row) < 13 or row[0] is None: 
            continue

        # Extraction des colonnes (basé sur votre structure)
        rubrique = str(row[0]).strip()       # Col 0
        nom = clean_text(str(row[2]) if row[2] else "")  # Col 2 (Libellé)
        seuil_alerte = row[7] if row[7] else 0    # Col 7
        image_url = row[9] if row[9] else ""      # Col 9
        prod_id = row[10] if row[10] else ""      # Col 10
        prix = row[12]                            # Col 12 (CAISSE)

        # Ignorer si pas de prix ou pas de nom
        if not prix or not nom:
            continue

        # Trouver le slug de la catégorie
        cat_slug = CATEGORY_MAPPING.get(rubrique)
        
        if cat_slug:
            slug = create_slug(nom, prod_id)
            
            # Création du JSON pour les traductions
            translations = json.dumps({
                "fr": {"name": nom.replace("''", "'")}, 
                "en": {"name": nom.replace("''", "'")}
            })

            # Requête SQL pour ce produit
            # On utilise une sous-requête pour trouver l'ID de la catégorie dynamiquement
            sql = f"""
            INSERT INTO products (category_id, slug, price, translations, image_url, low_stock_threshold, stock, is_active)
            VALUES (
                (SELECT id FROM categories WHERE slug = '{cat_slug}'),
                '{slug}',
                {prix},
                '{translations}',
                '{image_url}',
                {seuil_alerte},
                0, -- Stock par défaut à 0 car non fourni dans le CSV (sauf le seuil)
                true
            ) ON CONFLICT (slug) DO NOTHING;
            """
            sql_statements.append(sql)

except FileNotFoundError:
    print(f"Erreur : Le fichier {INPUT_FILE} est introuvable.")
except Exception as e:
    print(f"Erreur lors de la lecture du fichier Excel : {e}")

sql_statements.append("COMMIT;")

# Écriture du fichier SQL
with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_statements))

print(f"Succès ! Le fichier {OUTPUT_FILE} a été généré avec {len(sql_statements)} instructions.")